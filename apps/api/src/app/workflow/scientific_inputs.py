"""Resolve scientific task input refs against current persisted authorities."""

from __future__ import annotations

from base64 import b64encode
from collections.abc import Callable
import csv
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO, StringIO
import json
from uuid import UUID, uuid4

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import (
    ArtifactVersionModel,
    ResearchInputContentModel,
    ResearchInputModel,
    SourceSnapshotModel,
)
from app.schemas.core import ScientificSkillId, ScientificTaskInput
from app.schemas._hashing import compute_canonical_payload_hash
from app.schemas.data_artifacts import DatasetArtifactCandidate
from app.schemas.scientific_skills import ModelArtifactContent
from app.services.content_storage import ContentStorage, sha256_content_hash
from app.services.image_dataset import (
    ImageDatasetPolicy,
    resolve_image_dataset_archive,
)
from services.scientific_skills.execution import ScientificInputBinding
from services.scientific_skills.types import ScientificSourceReference


_ROW_SKILLS = frozenset(
    {
        ScientificSkillId.data_profile,
        ScientificSkillId.statistical_analysis,
        ScientificSkillId.correlation_analysis,
        ScientificSkillId.clustering_analysis,
        ScientificSkillId.anomaly_detection,
        ScientificSkillId.chart_visualization,
        ScientificSkillId.tabular_machine_learning,
        ScientificSkillId.time_series_classification,
        ScientificSkillId.time_series_forecast,
        ScientificSkillId.model_inference,
    }
)


class DatabaseScientificInputResolver:
    """Project-scoped resolver for ArtifactVersion, SourceSnapshot and input refs."""

    def __init__(
        self,
        session_factory: Callable[[], Session],
        content_storage: ContentStorage,
        *,
        project_id: str,
        image_dataset_policy: ImageDatasetPolicy | None = None,
    ) -> None:
        self._session_factory = session_factory
        self._content_storage = content_storage
        self._project_id = UUID(project_id)
        self._image_dataset_policy = image_dataset_policy or ImageDatasetPolicy()

    async def resolve(
        self, task: ScientificTaskInput
    ) -> tuple[ScientificInputBinding, ...]:
        return tuple(
            [await self._resolve_one(task, ref_id) for ref_id in task.input_refs]
        )

    async def _resolve_one(
        self, task: ScientificTaskInput, ref_id: str
    ) -> ScientificInputBinding:
        try:
            reference_uuid = UUID(ref_id)
        except ValueError as exc:
            raise ValueError(
                "scientific input refs must use persisted UUID identities"
            ) from exc
        with self._session_factory() as session, session.begin():
            version = session.scalar(
                select(ArtifactVersionModel).where(
                    ArtifactVersionModel.id == reference_uuid,
                    ArtifactVersionModel.project_id == self._project_id,
                )
            )
            if version is not None:
                return await _artifact_binding(
                    session,
                    task,
                    version,
                    self._content_storage,
                )
            source = session.scalar(
                select(SourceSnapshotModel).where(
                    SourceSnapshotModel.id == reference_uuid,
                    SourceSnapshotModel.project_id == self._project_id,
                )
            )
            if source is not None:
                return _source_binding(source)
            research_input = session.scalar(
                select(ResearchInputModel).where(
                    ResearchInputModel.id == reference_uuid,
                    ResearchInputModel.project_id == self._project_id,
                    ResearchInputModel.expires_at.is_(None),
                ).with_for_update()
            )
            if research_input is None:
                raise ValueError(
                    "scientific input ref was not found in the Run Project"
                )
            content_hash = research_input.content_hash
            input_type = research_input.type
            source_snapshot_id = research_input.source_snapshot_id
            source_reference = None
            if input_type == "image_dataset":
                source_reference = _image_dataset_source_reference(
                    session,
                    research_input,
                    project_id=self._project_id,
                )
            elif source_snapshot_id is not None:
                snapshot = session.scalar(
                    select(SourceSnapshotModel).where(
                        SourceSnapshotModel.id == source_snapshot_id,
                        SourceSnapshotModel.project_id == self._project_id,
                    )
                )
                if snapshot is None:
                    raise ValueError("Research Input SourceSnapshot is missing")
                source_reference = ScientificSourceReference(
                    source_snapshot_id=str(snapshot.id),
                    content_hash=snapshot.content_hash,
                )
        content = await self._content_storage.retrieve(content_hash)
        if content is None:
            raise ValueError("scientific input content blob is missing")
        return ScientificInputBinding(
            ref_id=ref_id,
            kind="content_blob",
            parameters=_content_parameters(
                task.skill_id,
                content,
                input_type=input_type,
                image_dataset_policy=self._image_dataset_policy,
            ),
            source_references=(source_reference,)
            if source_reference is not None
            else (),
        )


async def _artifact_binding(
    session: Session,
    task: ScientificTaskInput,
    version: ArtifactVersionModel,
    content_storage: ContentStorage,
) -> ScientificInputBinding:
    if task.skill_id not in _ROW_SKILLS:
        raise ValueError(
            f"{task.skill_id.value} does not accept an ArtifactVersion input"
        )
    if (
        task.skill_id is ScientificSkillId.model_inference
        and version.content.get("kind") == "model_artifact"
    ):
        model = ModelArtifactContent.model_validate(version.content)
        if model.status != "active":
            raise ValueError("model inference requires an active ModelArtifact")
        binary = await content_storage.retrieve(model.model_binary.content_hash)
        if binary is None:
            raise ValueError("ModelArtifact ONNX content is missing")
        if sha256_content_hash(binary) != model.model_binary.content_hash:
            raise ValueError("ModelArtifact ONNX content hash is invalid")
        snapshots = _source_references(
            session,
            project_id=version.project_id,
            source_snapshot_ids=tuple(version.source_snapshot_ids),
        )
        return ScientificInputBinding(
            ref_id=str(version.id),
            kind="artifact_version",
            parameters={
                "model": {
                    "model_artifact_version_id": str(version.id),
                    "model_id": model.model_id,
                    "task_kind": model.task_kind.value,
                    "feature_fields": list(model.feature_fields),
                    "target_field": model.target_field,
                    "content_base64": b64encode(binary).decode("ascii"),
                    "content_hash": model.model_binary.content_hash,
                    "media_type": model.model_binary.media_type,
                    "input_name": model.input_name,
                    "output_names": list(model.output_names),
                    "input_shape": list(model.input_shape),
                    "opset_imports": dict(model.opset_imports),
                }
            },
            source_references=snapshots,
            evidence_ids=tuple(sorted(version.evidence_ids)),
        )
    if version.content.get("kind") != "dataset":
        raise ValueError(f"{task.skill_id.value} requires a Dataset ArtifactVersion")
    dataset = DatasetArtifactCandidate.model_validate(version.content)
    snapshots = _source_references(
        session,
        project_id=version.project_id,
        source_snapshot_ids=tuple(version.source_snapshot_ids),
    )
    return ScientificInputBinding(
        ref_id=str(version.id),
        kind="artifact_version",
        parameters={"rows": _dataset_rows(dataset)}
        | (
            {"dataset_artifact_version_id": str(version.id)}
            if task.skill_id is ScientificSkillId.model_inference
            else {}
        ),
        source_references=snapshots,
        evidence_ids=tuple(sorted(version.evidence_ids)),
    )


def _source_binding(source: SourceSnapshotModel) -> ScientificInputBinding:
    return ScientificInputBinding(
        ref_id=str(source.id),
        kind="source_snapshot",
        parameters={},
        source_references=(
            ScientificSourceReference(
                source_snapshot_id=str(source.id),
                content_hash=source.content_hash,
            ),
        ),
    )


def _image_dataset_source_reference(
    session: Session,
    research_input: ResearchInputModel,
    *,
    project_id: UUID,
) -> ScientificSourceReference:
    if research_input.source_type != "upload" or research_input.status != "accepted":
        raise ValueError("image dataset requires an accepted upload Research Input")
    content = session.get(
        ResearchInputContentModel,
        (project_id, research_input.content_hash),
    )
    if content is None or content.mime_type != "application/zip":
        raise ValueError("image dataset Research Input MIME provenance is invalid")
    snapshot = None
    if research_input.source_snapshot_id is not None:
        snapshot = session.scalar(
            select(SourceSnapshotModel).where(
                SourceSnapshotModel.id == research_input.source_snapshot_id,
                SourceSnapshotModel.project_id == project_id,
            )
        )
        if snapshot is None:
            raise ValueError("image dataset SourceSnapshot reference is dangling")
    else:
        source_id = f"research_input:{research_input.id}"
        snapshot = session.scalar(
            select(SourceSnapshotModel).where(
                SourceSnapshotModel.project_id == project_id,
                SourceSnapshotModel.source_id == source_id,
                SourceSnapshotModel.source_type == "research_input_upload",
                SourceSnapshotModel.content_hash == research_input.content_hash,
            )
        )
        if snapshot is None:
            query = {"research_input_id": str(research_input.id)}
            snapshot = SourceSnapshotModel(
                id=uuid4(),
                project_id=project_id,
                source_id=source_id,
                source_type="research_input_upload",
                retrieved_at=research_input.created_at,
                query=query,
                query_hash=compute_canonical_payload_hash(query),
                source_version_or_etag=None,
                content_hash=research_input.content_hash,
                license_note="user-provided upload",
                cache_version=None,
                request_metadata={"ingestion_source": "upload"},
            )
            session.add(snapshot)
            session.flush()
        research_input.source_snapshot_id = snapshot.id
    if (
        snapshot.project_id != project_id
        or snapshot.source_id != f"research_input:{research_input.id}"
        or snapshot.source_type != "research_input_upload"
        or snapshot.content_hash != research_input.content_hash
    ):
        raise ValueError("image dataset SourceSnapshot provenance is invalid")
    return ScientificSourceReference(
        source_snapshot_id=str(snapshot.id),
        content_hash=snapshot.content_hash,
    )


def _source_references(
    session: Session,
    *,
    project_id: UUID,
    source_snapshot_ids: tuple[str, ...],
) -> tuple[ScientificSourceReference, ...]:
    if not source_snapshot_ids:
        return ()
    try:
        ids = tuple(UUID(item) for item in source_snapshot_ids)
    except ValueError as exc:
        raise ValueError("ArtifactVersion SourceSnapshot ids are invalid") from exc
    rows = tuple(
        session.scalars(
            select(SourceSnapshotModel).where(
                SourceSnapshotModel.id.in_(ids),
                SourceSnapshotModel.project_id == project_id,
            )
        )
    )
    by_id = {row.id: row for row in rows}
    if set(by_id) != set(ids):
        raise ValueError("ArtifactVersion SourceSnapshot closure is incomplete")
    return tuple(
        ScientificSourceReference(
            source_snapshot_id=str(item),
            content_hash=by_id[item].content_hash,
        )
        for item in ids
    )


def _dataset_rows(dataset: DatasetArtifactCandidate) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in dataset.rows:
        projection: dict[str, object] = {"row_id": row.row_id}
        for field in row.fields:
            if field.status == "mapped":
                projection[field.canonical_field_id] = _coerce_scalar(
                    field.canonical_value
                )
            else:
                projection[field.canonical_field_id] = None
        rows.append(projection)
    return rows


def _coerce_scalar(value: str) -> object:
    normalized = value.strip()
    try:
        return int(normalized)
    except ValueError:
        try:
            return float(normalized)
        except ValueError:
            return normalized


def _content_parameters(
    skill_id: ScientificSkillId,
    content: bytes,
    *,
    input_type: str,
    image_dataset_policy: ImageDatasetPolicy | None = None,
) -> dict[str, object]:
    if skill_id in {
        ScientificSkillId.ephemeris,
        ScientificSkillId.celestial_events,
    }:
        return {"ephemeris_base64": b64encode(content).decode("ascii")}
    if skill_id is ScientificSkillId.fits_image_analysis:
        if input_type != "fits":
            raise ValueError("fits_image_analysis requires a FITS Research Input")
        return {"fits_base64": b64encode(content).decode("ascii")}
    if skill_id is ScientificSkillId.image_classification:
        if input_type != "image_dataset":
            raise ValueError(
                "image_classification requires an image_dataset Research Input"
            )
        return resolve_image_dataset_archive(
            content,
            policy=image_dataset_policy or ImageDatasetPolicy(),
        )
    if skill_id in _ROW_SKILLS:
        if input_type == "csv":
            return {"rows": _csv_rows(content)}
        if input_type == "xlsx":
            return {"rows": _xlsx_rows(content)}
        if input_type == "parquet":
            return {"rows": _parquet_rows(content)}
        if input_type == "json":
            return {"rows": _json_rows(content)}
        raise ValueError(
            f"{skill_id.value} requires a CSV, XLSX, Parquet or JSON Research Input"
        )
    if skill_id in {
        ScientificSkillId.spectrum_analysis,
        ScientificSkillId.light_curve_analysis,
    }:
        if input_type != "json":
            raise ValueError(f"{skill_id.value} requires a JSON Research Input")
        payload = _json_object(content)
        return payload
    raise ValueError(f"{skill_id.value} does not accept a content blob input")


def _csv_rows(content: bytes) -> list[dict[str, object]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV Research Input must use UTF-8 encoding") from exc
    try:
        reader = csv.DictReader(StringIO(text), strict=True)
        if reader.fieldnames is None:
            raise ValueError("CSV Research Input requires a header row")
        fields = tuple(field.strip() for field in reader.fieldnames)
        if (
            not fields
            or len(fields) > 256
            or any(not field for field in fields)
            or len(fields) != len(set(fields))
        ):
            raise ValueError("CSV header must contain 1-256 unique non-empty fields")
        rows: list[dict[str, object]] = []
        for index, raw in enumerate(reader):
            if index >= 10_000:
                raise ValueError("CSV Research Input exceeds the 10000 row limit")
            if None in raw:
                raise ValueError("CSV row contains more cells than the header")
            rows.append(
                {
                    field: _coerce_uploaded_scalar(raw.get(original))
                    for field, original in zip(fields, reader.fieldnames, strict=True)
                }
            )
    except csv.Error as exc:
        raise ValueError("CSV Research Input is malformed") from exc
    if not rows:
        raise ValueError("CSV Research Input contains no data rows")
    return rows


def _json_rows(content: bytes) -> list[dict[str, object]]:
    payload = _decode_json(content)
    raw_rows = payload.get("rows") if isinstance(payload, dict) else payload
    if not isinstance(raw_rows, list) or not raw_rows:
        raise ValueError("JSON Research Input requires a non-empty rows array")
    if len(raw_rows) > 10_000:
        raise ValueError("JSON Research Input exceeds the 10000 row limit")
    rows: list[dict[str, object]] = []
    for raw in raw_rows:
        if (
            not isinstance(raw, dict)
            or len(raw) > 256
            or any(not isinstance(key, str) or not key.strip() for key in raw)
        ):
            raise ValueError("each JSON row must use 1-256 named fields")
        rows.append(dict(raw))
    return rows


def _xlsx_rows(content: bytes) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001 - dependency exceptions are not stable
        raise ValueError("XLSX Research Input is malformed") from exc
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError("XLSX Research Input requires exactly one worksheet")
        worksheet = workbook[workbook.sheetnames[0]]
        if worksheet.max_column > 256 or worksheet.max_row > 10_001:
            raise ValueError("XLSX Research Input exceeds the row or field limit")
        iterator = worksheet.iter_rows(values_only=True)
        try:
            raw_header = next(iterator)
        except StopIteration as exc:
            raise ValueError("XLSX Research Input requires a header row") from exc
        fields = _tabular_fields(raw_header, format_name="XLSX")
        rows: list[dict[str, object]] = []
        for raw_row in iterator:
            values = tuple(raw_row[: len(fields)])
            if any(value is not None for value in raw_row[len(fields) :]):
                raise ValueError("XLSX row contains more cells than the header")
            if all(value is None for value in values):
                continue
            rows.append(
                {
                    field: _normalize_tabular_scalar(value, format_name="XLSX")
                    for field, value in zip(fields, values, strict=True)
                }
            )
            if len(rows) > 10_000:
                raise ValueError("XLSX Research Input exceeds the 10000 row limit")
    finally:
        workbook.close()
    if not rows:
        raise ValueError("XLSX Research Input contains no data rows")
    return rows


def _parquet_rows(content: bytes) -> list[dict[str, object]]:
    import pyarrow as pa
    import pyarrow.parquet as pq

    try:
        parquet = pq.ParquetFile(BytesIO(content))
        metadata = parquet.metadata
        if metadata.num_rows <= 0 or metadata.num_rows > 10_000:
            raise ValueError("Parquet Research Input exceeds the 10000 row limit")
        if metadata.num_columns <= 0 or metadata.num_columns > 256:
            raise ValueError("Parquet Research Input exceeds the 256 field limit")
        if (
            sum(
                metadata.row_group(index).total_byte_size
                for index in range(metadata.num_row_groups)
            )
            > 100 * 1024 * 1024
        ):
            raise ValueError("Parquet Research Input exceeds the decoded byte limit")
        fields = _tabular_fields(parquet.schema_arrow.names, format_name="Parquet")
        if any(not _supported_arrow_type(field.type, pa) for field in parquet.schema_arrow):
            raise ValueError("Parquet Research Input contains a nested or binary field")
        raw_rows = parquet.read(use_threads=False).to_pylist()
    except ValueError:
        raise
    except Exception as exc:  # noqa: BLE001 - dependency exceptions are not stable
        raise ValueError("Parquet Research Input is malformed") from exc
    return [
        {
            field: _normalize_tabular_scalar(raw[field], format_name="Parquet")
            for field in fields
        }
        for raw in raw_rows
    ]


def _tabular_fields(
    raw_fields: object,
    *,
    format_name: str,
) -> tuple[str, ...]:
    if not isinstance(raw_fields, (list, tuple)):
        raw_fields = tuple(raw_fields)  # type: ignore[arg-type]
    fields = tuple(
        value.strip() if isinstance(value, str) else "" for value in raw_fields
    )
    if (
        not fields
        or len(fields) > 256
        or any(not field for field in fields)
        or len(fields) != len(set(fields))
    ):
        raise ValueError(
            f"{format_name} header must contain 1-256 unique non-empty fields"
        )
    return fields


def _supported_arrow_type(value: object, pa: object) -> bool:
    types = pa.types  # type: ignore[attr-defined]
    return any(
        predicate(value)
        for predicate in (
            types.is_null,
            types.is_boolean,
            types.is_integer,
            types.is_floating,
            types.is_decimal,
            types.is_string,
            types.is_large_string,
            types.is_date,
            types.is_time,
            types.is_timestamp,
        )
    )


def _normalize_tabular_scalar(value: object, *, format_name: str) -> object:
    if value is None or isinstance(value, str | bool | int | float):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    raise ValueError(f"{format_name} Research Input contains an unsupported scalar")


def _json_object(content: bytes) -> dict[str, object]:
    payload = _decode_json(content)
    if not isinstance(payload, dict) or len(payload) > 64:
        raise ValueError("scientific JSON input must be a bounded object")
    return dict(payload)


def _decode_json(content: bytes) -> object:
    try:
        return json.loads(content)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise ValueError("JSON Research Input must be valid UTF-8 JSON") from exc


def _coerce_uploaded_scalar(value: str | None) -> object:
    if value is None or not value.strip():
        return None
    return _coerce_scalar(value)


__all__ = ["DatabaseScientificInputResolver"]
